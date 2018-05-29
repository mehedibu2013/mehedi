import selenium
import openpyxl
import os
from selenium import webdriver
# from selenium.webdriver.support.select import Select
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.chrome.options import Options


# Chrome driver
options = webdriver.ChromeOptions()
options.add_argument("--disable-infobars")
options.add_argument("test-type")
options.add_argument('--ignore-certificate-errors')
options.add_argument("disable-extensions")
options.add_argument("--start-maximized")
options.add_argument("--enable-automation")
# browser= webdriver.Chrome("C:\Program Files\Google\Chrome\Application\chromedriver.exe",chrome_options=options)
time.sleep(2)

# IE driver
# browser = webdriver.Ie("C:\IEDriverServer.exe")
'''
browser.get("http://10.16.41.84:8091/pmi/login.jsp")
time.sleep(2)
user=browser.find_element_by_id("j_username")
user.send_keys("nafi3496")
time.sleep(2)
password=browser.find_element_by_id("j_password")
password.clear()
password.send_keys("")
loginPMI=browser.find_element_by_name("login")
loginPMI.click()
browser.find_element_by_link_text("Policy").click()
'''

# read and create data object
book = openpyxl.load_workbook("pmi_data_LIVE.xlsx")
sheet = book['Sheet1']
time.sleep(2)

for i in range(2,3):

    # SERVICE DEFINITION
    browser = webdriver.Chrome("E:\chromedriver.exe", chrome_options=options)
    browser.get("http://10.16.41.71:8091/pmi/login.jsp")
    time.sleep(2)
    user = browser.find_element_by_id("j_username")
    user.send_keys("")
    time.sleep(2)
    password = browser.find_element_by_id("j_password")
    password.clear()
    password.send_keys("")
    loginPMI = browser.find_element_by_name("login")
    loginPMI.click()
    browser.find_element_by_link_text("Policy").click()

    time.sleep(2)
    browser.find_element_by_link_text("PCRF").click()
    time.sleep(2)
    browser.find_element_by_link_text("Service Flows").click()
    time.sleep(2)
    browser.find_element_by_xpath('//*[@id="tabContainer_tablist_predefinedTab"]/span[1]').click()
    browser.find_element_by_xpath('//*[@id="addServiceRule"]').click()
    time.sleep(3)
    browser.switch_to.window(browser.current_window_handle)
    ServiceName = browser.find_element_by_id("name")
    time.sleep(1)
    ServiceName.send_keys(sheet.cell(row=2, column=i).value)
    MK = browser.find_element_by_id("monitorKey")
    MK.send_keys(sheet.cell(row=3, column=i).value)
    RN = browser.find_element_by_id("ruleName")
    RN.send_keys(sheet.cell(row=2, column=i).value)
    RP = browser.find_element_by_id("rulePriority")
    RP.send_keys(sheet.cell(row=5, column=i).value)
    browser.find_element_by_id("submitServiceRule").click()
    time.sleep(3)

    browser.switch_to.window(browser.current_window_handle)
    browser.find_element_by_id("addServiceProfile_label").click()
    browser.switch_to.window(browser.current_window_handle)
    time.sleep(3)

    # Serive profile definition
    SPNAME = "SP_{}".format(sheet.cell(row=2, column=i).value)
    browser.find_element_by_id("name").send_keys(SPNAME)
    SPTYPE = browser.find_element_by_id("serviceProfileType")
    SPTYPE.clear()
    SPTYPE.send_keys("PREDEFINED")
    SerRestrict = browser.find_element_by_id("serviceRestriction")
    SerRestrict.clear()
    SerRestrict.send_keys("NONE")
    SerRule = browser.find_element_by_id("serviceRule")
    SerRule.send_keys(sheet.cell(row=2, column=i).value)
    browser.find_element_by_id("saveServiceProfile_label").click()

    # Plan definition
    time.sleep(5)
   # browser.find_element_by_link_text("Policy").click()
    browser.find_element_by_link_text("SPCM").click()
    time.sleep(2)
    browser.find_element_by_link_text("Plan Definitions").click()
    time.sleep(5)
    browser.find_element_by_link_text("Subscriber Plans").click()
    time.sleep(2)
    browser.find_element_by_xpath("//span[contains(text(), 'In Progress')]").click()
    time.sleep(1)
    browser.find_element_by_id("add_label").click()
    time.sleep(5)
    browser.find_element_by_name("planDefinition.name").send_keys(sheet.cell(row=6, column=i).value)
    browser.find_element_by_name("planDefinition.description").send_keys(sheet.cell(row=7, column=i).value)

    if sheet.cell(row=8, column=i).value == "Yes":
        browser.find_element_by_name("planDefinition.coreUpdate").click()

    if sheet.cell(row=9, column=i).value != "Yes":
        browser.find_element_by_name("planDefinition.allowed").click()

    if sheet.cell(row=10, column=i).value == "Yes":
        browser.find_element_by_id("chargable").click()
        browser.find_element_by_id("cost").send_keys(sheet.cell(row=11, column=i).value)
    browser.find_element_by_id("maxDeactivationCount").send_keys(sheet.cell(row=12, column=i).value)
    browser.find_element_by_id("maxDeactivationPeriod").send_keys(sheet.cell(row=13, column=i).value)
    browser.find_element_by_id("planPrecedence").clear()
    browser.find_element_by_id("planPrecedence").send_keys(sheet.cell(row=14, column=i).value)
    browser.find_element_by_id("displayUnitAmountVolume").send_keys(sheet.cell(row=15, column=i).value)
    browser.find_element_by_id("validityPeriodCheckbox").click()
    browser.find_element_by_id("validityPeriod").send_keys(sheet.cell(row=16, column=i).value)
    time.sleep(2)
    if sheet.cell(row=17, column=i).value != "24hours":
        browser.find_element_by_id("absoluteExpiryTimeCheckbox").click()
        time.sleep(2)
        browser.find_element_by_id("absoluteExpiryTime").send_keys(sheet.cell(row=17, column=i).value)
    if sheet.cell(row=18, column=i).value == "Yes":
        browser.find_element_by_id("accumulationPermittedCheckbox").click()
        time.sleep(3)
        browser.find_element_by_id("displayAccumulationRollOverLimitVolume").clear()
        browser.find_element_by_id("displayAccumulationRollOverLimitVolume").send_keys(sheet.cell(row=19, column=i).value)

    # Plan purchase restriction
    if sheet.cell(row=20, column=i).value != "NA":
        browser.find_element_by_id("maxCount").send_keys(sheet.cell(row=20, column=i).value)
        time.sleep(2)
        browser.find_element_by_name("purchaseRestriction.period").send_keys(sheet.cell(row=21, column=i).value)
        browser.find_element_by_id("periodType").clear()
        time.sleep(1)
        browser.find_element_by_id("periodType").send_keys(sheet.cell(row=22, column=i).value)
    time.sleep(2)

    # Renewal count/rollover
    if sheet.cell(row=23, column=i).value != "NA":
        browser.find_element_by_id("recurring").click()
        browser.find_element_by_id("maxOccurrenceCount").send_keys(str(sheet.cell(row=24, column=i).value))
        browser.find_element_by_id("maxParkingRetryCount").send_keys(sheet.cell(row=25, column=i).value)
        browser.find_element_by_name("planDefinition.parkingRetryDelay").send_keys(sheet.cell(row=26, column=i).value)
        time.sleep(2)
        ''''
        # Renewal rollover
        browser.find_element_by_id("recycleRollOverLimitCheckbox").click()
        browser.find_element_by_id("displayRecycleRollOverLimitVolume").clear()
        browser.find_element_by_id("displayRecycleRollOverLimitVolume").send_keys(sheet.cell(row=27, column=i).value)
    time.sleep(2)'''

    # Identifier
    if sheet.cell(row=28, column=i).value != "NA":
        browser.find_element_by_id("identifier").send_keys(sheet.cell(row=28, column=i).value)
    time.sleep(4)

    # PCC PROFILE
    browser.find_element_by_id("addPccProfileButton_label").click()
    time.sleep(2)
    browser.switch_to.window(browser.current_window_handle)
    PredefinedRuleName = sheet.cell(row=2, column=i).value
    PCC = PredefinedRuleName + "_PCC"
    time.sleep(2)
    browser.find_element_by_id("alias").send_keys(PCC)
    browser.find_element_by_id("profile.serviceProfile").clear()
    SPNAME = "SP_{}".format(sheet.cell(row=2, column=i).value)
    browser.find_element_by_id("profile.serviceProfile").send_keys(SPNAME)
    browser.find_element_by_id("profile.chargingProfile").clear()
    time.sleep(1)
    browser.find_element_by_id("profile.chargingProfile").send_keys("Prepaid Postpaid Bundle")
    browser.find_element_by_id("standardLocationProfiles").clear()
    time.sleep(1)
    browser.find_element_by_id("standardLocationProfiles").send_keys("HOME-PLMNID")
    time.sleep(2)

    if sheet.cell(row=29, column=i).value != "24hours":
        browser.find_element_by_id("profile.timeProfile").clear()
        time.sleep(2)
        browser.find_element_by_id("profile.timeProfile").send_keys(sheet.cell(row=29, column=i).value)
    else:
        browser.find_element_by_id("profile.timeProfile").clear()
        time.sleep(2)
        browser.find_element_by_id("profile.timeProfile").send_keys("None")


    #browser.find_element_by_id("savePccProfile_label").click()
    browser.find_element_by_id("profile.timeProfile").send_keys(Keys.TAB)
    time.sleep(2)
   # browser.find_element_by_id("profile.networkProfile").send_keys(Keys.TAB)
    time.sleep(2)

    if sheet.cell(row=33, column=i).value == "RAT_4G":
        browser.find_element_by_id("profile.networkProfile").clear()
        time.sleep(2)
        browser.find_element_by_id("profile.networkProfile").send_keys(sheet.cell(row=33, column=i).value)
    else:
        browser.find_element_by_id("profile.networkProfile").clear()
        time.sleep(2)
        browser.find_element_by_id("profile.networkProfile").send_keys("None")
        time.sleep(2)

    browser.find_element_by_id("profile.networkProfile").send_keys(Keys.TAB)
    time.sleep(2)
    browser.find_element_by_id("profile.deviceProfile").send_keys(Keys.TAB)
    time.sleep(1)
    browser.find_element_by_id("profile.stateProfile").send_keys(Keys.TAB)
    time.sleep(1)
    browser.find_element_by_id("meteringPercentage").send_keys(Keys.TAB)
    time.sleep(2)
    #browser.find_element_by_id("savePccProfile_label").send_keys(Keys.ENTER)
    browser.find_element_by_id("savePccProfile_label").click()

    # USAGE COUNTER
    time.sleep(4)
    browser.switch_to.window(browser.current_window_handle)
    time.sleep(1)
    browser.find_element_by_id("addUsageCounterButton_label").click()
    browser.switch_to.window(browser.current_window_handle)
    time.sleep(2)
    UsageCounter = sheet.cell(row=2, column=i).value
    UC = UsageCounter + "_UC"
    browser.find_element_by_id("usageCounterDefinition.name").send_keys(UC)
    browser.find_element_by_id("usageCounterDefinition.resetTimeType").clear()
    time.sleep(2)
    browser.find_element_by_id("usageCounterDefinition.resetTimeType").send_keys("Relative")
    time.sleep(4)
    browser.find_element_by_id("usageCounterDefinition.timeUnit").clear()
    browser.find_element_by_id("usageCounterDefinition.timeUnit").send_keys("None")
    browser.find_element_by_id("saveUsageCounter").click()

    # USAGE REPORT
    time.sleep(5)
    browser.switch_to.window(browser.current_window_handle)
    time.sleep(5)
    browser.find_element_by_id("displayDefaultGrantedVolumeAmount").clear()
    time.sleep(2)
    browser.find_element_by_id("displayDefaultGrantedVolumeAmount").send_keys(sheet.cell(row=30, column=i).value)

    # USAGE RULE
    time.sleep(3)
    browser.find_element_by_id("addUsageRuleButton_label").click()
    time.sleep(2)
    browser.switch_to.window(browser.current_window_handle)
    browser.find_element_by_id("usageRuleDefinition.name").send_keys("50pct usage notification")
    browser.find_element_by_id("displayThresholdVolume").clear()
    time.sleep(4)
    browser.find_element_by_id("displayThresholdVolume").send_keys(sheet.cell(row=31, column=i).value)
    browser.find_element_by_id("usageRuleDefinition.action").clear()
    time.sleep(2)
    browser.find_element_by_id("usageRuleDefinition.action").send_keys("AT_MSG_Plan_50_Consumed")
    time.sleep(2)
    browser.find_element_by_id("usageRuleDefinition.action").send_keys(Keys.TAB)
    browser.find_element_by_id("createAction").send_keys(Keys.TAB)
    time.sleep(2)
    browser.find_element_by_id("saveUsageRule").send_keys(Keys.ENTER)

    time.sleep(3)
    browser.switch_to.window(browser.current_window_handle)
    browser.find_element_by_id("addUsageRuleButton_label").click()
    time.sleep(2)
    browser.switch_to.window(browser.current_window_handle)
    browser.find_element_by_id("usageRuleDefinition.name").send_keys("80pct usage notification")
    browser.find_element_by_id("displayThresholdVolume").send_keys(sheet.cell(row=32, column=i).value)
    browser.find_element_by_id("usageRuleDefinition.action").clear()
    time.sleep(2)
    browser.find_element_by_id("usageRuleDefinition.action").send_keys("AT_MSG_Plan_80_Consumed")
    time.sleep(2)
    browser.find_element_by_id("usageRuleDefinition.action").send_keys(Keys.TAB)
    time.sleep(1)
    browser.find_element_by_id("createAction").send_keys(Keys.TAB)
    time.sleep(2)
    browser.find_element_by_id("saveUsageRule").send_keys(Keys.ENTER)
    browser.switch_to.window(browser.current_window_handle)
    time.sleep(2)
    browser.find_element_by_id("save_label").click()
    time.sleep(10)
    browser.close()